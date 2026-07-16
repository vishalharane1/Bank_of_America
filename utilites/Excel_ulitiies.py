import openpyxl

class Excel_Ulitiles_class:

    @staticmethod
    def get_max_row_from_excel(filepath,sheetname):
        excel_path=openpyxl.load_workbook(filepath)
        sheet=excel_path[sheetname]
        return sheet.max_row

    @staticmethod
    def read_data_from_excel(filepath,sheetname,rowno,columnno):
        excel_path=openpyxl.load_workbook(filepath)
        sheet=excel_path[sheetname]
        return sheet.cell(rowno,columnno).value

    @staticmethod
    def write_date_to_excel(filepath,sheetname,rowno,columnno,data):
        excelpath=openpyxl.load_workbook(filepath)
        sheet=excelpath[sheetname]
        sheet.cell(rowno,columnno).value=data
        excelpath.save(filepath)
        excelpath.close()
